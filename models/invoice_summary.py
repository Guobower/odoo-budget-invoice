# -*- coding: utf-8 -*-
# TODO replace sudo with suspend_security using OCA base_suspend_security
import tempfile
from openpyxl.drawing.image import Image

from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
from odoo.addons.budget_utilities.models.utilities import choices_tuple
from ..xlsx_creator.creator import Creator
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

from .invoice import Invoice


def format_footer(ws, row, sr):
    footer_cell = ws.cell(row=row + sr - 2, column=1)  # A15 + sr(row) - 2 row
    footer_cell.font = Font(size=11, bold=True)
    ws.row_dimensions[footer_cell.row].height = 83


def inject_signature(ws, signatory_img, signature_coor, sr):
    tmpfile = tempfile.TemporaryFile()
    signatory_img.save(tmpfile, format="PNG")
    ws.add_image(Image(tmpfile), "%s" % signature_coor[0] + str(int(signature_coor[1:]) + sr))


def inject_form_header(ws, creator, logo_coor, header_coor):
    fill = PatternFill(start_color='ADE75F', end_color='ADE75F', fill_type='solid')

    h_cell = ws.cell(header_coor)
    h_cell.fill = fill
    ws.add_image(creator.logo, logo_coor)
    return ws


def create_allocation_sheet(summary=None, wb=None):
    if summary is None and wb is None:
        raise NotImplementedError

    # CEAR ALLOCATIONS
    row = 2
    column = 1
    sr = 1
    ws = wb.create_sheet("CEAR ALLOCATION")
    cear_allocation_ids = summary.mapped('invoice_ids.cear_allocation_ids')

    # HEADING
    ws.cell(row=row - 1, column=column).value = 'SR'
    ws.cell(row=row - 1, column=column + 1).value = 'Invoice No'
    ws.cell(row=row - 1, column=column + 2).value = 'Invoice Certified Amount'
    ws.cell(row=row - 1, column=column + 3).value = 'Cear No'
    ws.cell(row=row - 1, column=column + 4).value = 'Allocated Amount'

    for allocation in cear_allocation_ids:
        ws.cell(row=row, column=column).value = sr
        ws.cell(row=row, column=column + 1).value = allocation.invoice_id.invoice_no
        ws.cell(row=row, column=column + 2).value = allocation.invoice_id.certified_invoice_amount
        ws.cell(row=row, column=column + 3).value = allocation.cear_id.no
        ws.cell(row=row, column=column + 4).value = allocation.amount

        row += 1
        sr += 1


def get_joined_value(vals):
    vals = set(vals)
    new_vals = []
    for val in vals:
        if isinstance(val, bool) or val is False:
            continue
        elif isinstance(val, fields.Date):
            val = fields.Datetime.from_string(val).strftime('%d-%b-%Y')
        else:
            val = '{}'.format(val)

        new_vals.append(val)

    if len(new_vals) == 1 and isinstance(new_vals[0], float):
        return float(new_vals[0])

    return ', '.join(new_vals)


class InvoiceSummary(models.Model):
    _name = 'budget.invoice.invoice.summary'
    _rec_name = 'summary_no'
    _description = 'Invoice Summary'
    _order = 'id desc'
    _inherit = ['mail.thread']

    # CHOICES
    # ----------------------------------------------------------
    STATES = choices_tuple(['draft', 'file generated', 'sd signed', 'svp signed', 'cto signed',
                            'sent to finance', 'closed', 'cancelled'], is_sorted=False)
    SIGNATURES = [
        ('signature0001.png', 'SD/TBPC & SVP/MN & CTO'),
        ('signature0002.png', 'SD/TBPC & SVP/CSE & SVP/MN & CTO'),
        ('signature0003.png', 'SVP/ET & SD/TBPC & SVP/MN & FINANCE'),
        ('signature0004.png', 'SVP/ET & SD/TBPC & SVP/MN & CTO & FINANCE')
    ]
    FORMS = [
        ('form_a0001ver02.xlsx', 'Regional'),
        ('form_a0002ver02.xlsx', 'Regional - Volume Discount'),
        ('form_b0001ver05.xlsx', 'Resource'),
        ('form_c0001ver02.xlsx', 'Head Office'),
    ]

    OBJECTIVES = choices_tuple(['invoice certification', 'on hold certification'])

    # BASIC FIELDS
    # ----------------------------------------------------------
    active = fields.Boolean(default=True, help="Set active to false to hide the tax without removing it.")
    is_head_office = fields.Boolean(default=False)
    is_regional = fields.Boolean(default=False)

    state = fields.Selection(STATES, default='draft', track_visibility='onchange')
    # TODO CONSIDER MAKING SUMMARY NO AS NAME
    # TODO USE GET DEFAULT_SUMMARY IN CREATE
    summary_no = fields.Char(string='Summary No',
                             default=lambda self: self._get_default_summary_no())
    form = fields.Selection(FORMS)
    signature = fields.Selection(SIGNATURES)
    objective = fields.Selection(OBJECTIVES, default='invoice certification')
    # TODO DEPRECATE
    team = Invoice.team

    # TODO DEPRECATE
    signed_date = fields.Date(string='Signed Date')
    # --------------

    sd_signed_date = fields.Date(string='SD Signed Date', track_visibility='onchange')
    svp_signed_date = fields.Date(string='SVP Signed Date', track_visibility='onchange')
    cto_signed_date = fields.Date(string='CTO Signed Date', track_visibility='onchange')

    closed_date = fields.Date(string='Closed Date', track_visibility='onchange')
    sent_finance_date = fields.Date(string='Sent to Finance Date', track_visibility='onchange')
    sequence = fields.Integer(string='Sequence')

    # RELATIONSHIPS
    # ----------------------------------------------------------
    responsible_id = fields.Many2one('res.users', string='Responsible',
                                     default=lambda self: self.env.user.id)
    # TODO MUST NOT EDITABLE WHEN ON PROCESS, CHECK RULE ACCESS
    invoice_ids = fields.Many2many('budget.invoice.invoice',
                                   'budget_invoice_summary_invoice',
                                   'summary_id',
                                   'invoice_id')
    signature_ids = fields.Many2many('budget.signature.signatory',
                                     'budget_invoice_summary_signature',
                                     'summary_id',
                                     'signature_id')
    # TODO REMOVE
    section_id = fields.Many2one('res.partner', string='Section')

    # COMPUTE FIELDS
    # ----------------------------------------------------------
    invoice_count = fields.Integer(string='Invoice Count',
                                   compute='_compute_invoice_count',
                                   store=True)

    @api.onchange('objective')
    def _onchange_invoice_ids_filter(self):
        if self.objective == 'invoice certification':
            return {'domain': {'invoice_ids': [('state', '=', 'verified'),
                                               ('is_head_office', '=', self.is_head_office),
                                               ('is_regional', '=', self.is_regional)
                                               ]}}
        elif self.objective == 'on hold certification':
            return {'domain': {'invoice_ids': [('has_hold_amount', '=', True),
                                               ('is_head_office', '=', self.is_head_office),
                                               ('is_regional', '=', self.is_regional)
                                               ]}}


    @api.one
    @api.depends('invoice_ids')
    def _compute_invoice_count(self):
        self.invoice_count = len(self.invoice_ids)

    # CONSTRAINTS
    # ----------------------------------------------------------
    _sql_constraints = [
        (
            '_uniq_summary_no',
            'UNIQUE(summary_no)',
            'summary must be unqiue!',
        )
    ]

    @api.model
    def _get_default_summary_no(self):
        from datetime import datetime as dt
        from dateutil.relativedelta import relativedelta
        now = dt.utcnow()
        month = '{:%^b}'.format(now)
        year = '{:%y}'.format(now)
        start = dt(now.year, now.month, 1)
        end = dt(now.year, now.month, 1) + relativedelta(months=1)
        # end should be the first day next month to make time consider as 23:59:99
        domain = [('create_date', '>=', fields.Datetime.to_string(start)),
                  ('create_date', '<', fields.Datetime.to_string(end))]

        summaries = self.search(domain)
        if len(summaries) == 0:
            sr = 1
        else:
            sr = summaries[0].summary_no.split('-')[-1]
            sr = int(sr) + 1

        return 'IM-%s%s-%03d' % (month, year, sr)

    @api.one
    @api.returns('self', lambda value: value.id)
    def copy(self, default=None):
        default = dict(default or {})
        default.update(summary_no=self._get_default_summary_no())
        return super(InvoiceSummary, self).copy(default)

    @api.one
    def form_a0001ver02(self, creator):
        """
        GENERATE FORM ACCORDING TO form_a0001ver01
        """

        wb = creator.get_wb()

        # CREATE ALLOCATION SHEET
        # ----------------------------------------------------------
        create_allocation_sheet(self, wb)

        # WORK SHEET MAIN
        # ----------------------------------------------------------
        row = 13
        column = 1
        sr = 1
        signature_coor = "B18"
        logo_coor = "K1"
        header_coor = "A3"
        footer_row = 16
        ws = wb.get_sheet_by_name('main')

        ws.cell("C5").value = self.summary_no
        ws.cell("C6").value = fields.Datetime.from_string(self.create_date).strftime('%d-%b-%Y')
        currency = self.mapped("invoice_ids.currency_id")
        ws.cell("C7").value = currency[0].name if currency else ""

        # Create Table
        ws.insert_rows(row, len(self.invoice_ids) - 1)

        # No, Reg, Contractor, Invoice No, Contract, Revenue, OpEx, CapEx, Total Amt, Budget/Yr.
        # 1 , 2  , 3,        , 4         , 5  6    , 7      , 8   , 9    , 10       , 11
        for r in self.invoice_ids.sorted(key=lambda rec: rec.sequence):
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = r.region_id.alias.upper() or ''
            ws.cell(row=row, column=column + 2).value = r.contract_id.contractor_id.name or ''
            ws.cell(row=row, column=column + 3).value = r.invoice_no or ''
            ws.cell(row=row, column=column + 4).value = r.contract_id.no or ''
            ws.cell(row=row, column=column + 5).value = r.revenue_amount or ''
            ws.cell(row=row, column=column + 6).value = r.opex_amount or ''
            ws.cell(row=row, column=column + 7).value = r.capex_amount or ''
            ws.cell(row=row, column=column + 8).value = r.certified_invoice_amount or ''
            # ws.cell(row=row, column=column + 9).value = r.cear_allocation_ids.cear_id.im_utilized_amount
            ws.cell(row=row, column=column + 10).value = get_joined_value(r.mapped('cear_allocation_ids.cear_id.no'))
            row += 1
            sr += 1

        # INSERT HEADER LOGO AND SIGNATURE
        inject_form_header(ws, creator, logo_coor, header_coor)

        # FORMAT FOOTER
        format_footer(ws, footer_row, sr)

        # SIGNATURE
        inject_signature(ws, self.get_signatories(), signature_coor, sr)

        # SAVE FINAL ATTACHMENT
        creator.save()
        creator.attach(self.env)

    @api.one
    def form_a0002ver02(self, creator):
        """
        GENERATE FORM ACCORDING TO form_a0002ver01
        """

        wb = creator.get_wb()

        # CREATE ALLOCATION SHEET
        # ----------------------------------------------------------
        create_allocation_sheet(self, wb)

        # WORK SHEET MAIN
        # ----------------------------------------------------------
        row = 13
        column = 1
        sr = 1
        signature_coor = "C18"
        logo_coor = "M1"
        header_coor = "A3"
        footer_row = 16
        ws = wb.get_sheet_by_name('main')

        ws.cell("C5").value = self.summary_no
        ws.cell("C6").value = fields.Datetime.from_string(self.create_date).strftime('%d-%b-%Y')
        currency = self.mapped("invoice_ids.currency_id")
        ws.cell("C7").value = currency[0].name if currency else ""

        # Create Table
        ws.insert_rows(row, len(self.invoice_ids) - 1)
        # No, Reg, Contractor, Invoice No, Contract, Revenue, OpEx, CapEx, Total Amt, Budget/Yr.
        # 1 , 2  , 3,        , 4         , 5  6    , 7      , 8   , 9    , 10       , 11
        for r in self.mapped('invoice_ids').sorted(key=lambda rec: rec.sequence):
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = r.region_id.alias.upper() or ''
            ws.cell(row=row, column=column + 2).value = r.contract_id.contractor_id.name or ''
            ws.cell(row=row, column=column + 3).value = r.invoice_no or ''
            ws.cell(row=row, column=column + 4).value = r.contract_id.no or ''
            ws.cell(row=row, column=column + 5).value = r.revenue_amount or ''
            ws.cell(row=row, column=column + 6).value = r.opex_amount or ''
            ws.cell(row=row, column=column + 7).value = r.capex_amount or ''
            # ws.cell(row=row, column=column + 8).value = r.currency_id.name or ''
            ws.cell(row=row, column=column + 8).value = r.invoice_amount or ''
            # ws.cell(row=row, column=column + 9).value = r.cear_allocation_ids.cear_id.im_utilized_amount
            ws.cell(row=row, column=column + 9).value = '' if not r.discount_percentage else r.discount_percentage / 100
            ws.cell(row=row, column=column + 10).value = r.discount_amount or ''
            ws.cell(row=row, column=column + 11).value = r.certified_invoice_amount or ''
            ws.cell(row=row, column=column + 12).value = get_joined_value(r.mapped('cear_allocation_ids.cear_id.no'))
            row += 1
            sr += 1

        # INSERT HEADER LOGO AND SIGNATURE
        inject_form_header(ws, creator, logo_coor, header_coor)

        # FORMAT FOOTER
        format_footer(ws, footer_row, sr)

        # SIGNATURE
        inject_signature(ws, self.get_signatories(), signature_coor, sr)

        # SAVE FINAL ATTACHMENT
        creator.save()
        creator.attach(self.env)

    @api.one
    def form_b0001ver05(self, creator):
        """
        GENERATE FORM ACCORDING TO form_b0001ver04
        """

        wb = creator.get_wb()

        # CREATE ALLOCATION SHEET
        # ----------------------------------------------------------
        create_allocation_sheet(self, wb)

        # WORK SHEET MAIN
        # ----------------------------------------------------------
        row = 22
        column = 2
        sr = 1
        signature_coor = "E28"
        logo_coor = "R1"
        header_coor = "B3"
        footer_row = 26
        ws = wb.get_sheet_by_name('main')

        ws.cell("B11").value = self.summary_no
        ws.cell("B15").value = fields.Datetime.from_string(self.create_date).strftime('%d-%b-%Y')
        ws.cell("E15").value = get_joined_value(self.mapped('invoice_ids.approval_ref'))
        ws.cell("K11").value = get_joined_value(self.mapped('invoice_ids.contractor_id.name'))
        ws.cell("K15").value = get_joined_value(self.mapped('invoice_ids.contract_id.contract_ref'))
        ws.cell("N15").value = get_joined_value(self.mapped('invoice_ids.division_id.alias'))
        currency = self.mapped("invoice_ids.currency_id")
        ws.cell("B19").value = currency[0].name if currency else ""

        # Create Table
        ws.insert_rows(row, len(self.invoice_ids) - 1)

        # No, Reg, Contractor, Invoice No, Contract, Revenue, OpEx, CapEx, Total Amt, Budget/Yr.
        # 1 , 2  , 3,        , 4         , 5  6    , 7      , 8   , 9    , 10       , 11
        for r in self.mapped('invoice_ids').sorted(key=lambda rec: rec.sequence):
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = fields.Datetime.from_string(r.invoice_date).strftime(
                '%d-%b-%Y') or ''
            ws.cell(row=row, column=column + 2).value = r.invoice_no or ''
            ws.cell(row=row, column=column + 3).value = r.po_id.no or ''
            ws.cell(row=row, column=column + 4).value = ', '.join(
                [i or '' for i in r.mapped('cear_allocation_ids.cear_id.no')])

            cc_ac = []
            charge_account = ''
            for i in r.oear_allocation_ids:
                if i.cost_center_id.cost_center == '71101':
                    charge_account = i.cost_center_id.cost_center
                else:
                    cc_ac.append('{} {}'.format(i.cost_center_id.cost_center,
                                                i.account_code_id.account_code))

            ws.cell(row=row, column=column + 5).value = ','.join(cc_ac) if len(cc_ac) > 0 else ''

            ws.cell(row=row, column=column + 6).value = charge_account
            ws.cell(row=row, column=column + 7).value = r.description or ''
            ws.cell(row=row, column=column + 8).value = r.invoice_amount
            ws.cell(row=row, column=column + 9).value = r.other_deduction_amount
            ws.cell(row=row, column=column + 10).value = r.invoice_amount - r.other_deduction_amount
            ws.cell(row=row, column=column + 11).value = r.discount_amount
            ws.cell(row=row, column=column + 12).value = r.certified_invoice_amount
            ws.cell(row=row, column=column + 13).value = r.certified_capex_amount
            ws.cell(row=row, column=column + 14).value = r.certified_revenue_amount
            ws.cell(row=row, column=column + 15).value = r.certified_opex_amount
            ws.cell(row=row, column=column + 16).value = r.remark or ''

            ws.cell(row=row, column=column + 18).value = r.discount_amount or 0
            ws.cell(row=row, column=column + 19).value = r.discount_percentage or 0

            row += 1
            sr += 1

        # INSERT HEADER LOGO AND SIGNATURE
        inject_form_header(ws, creator, logo_coor, header_coor)

        # FORMAT FOOTER
        format_footer(ws, footer_row, sr)

        # SIGNATURE
        inject_signature(ws, self.get_signatories(), signature_coor, sr)

        # SAVE FINAL ATTACHMENT
        creator.save()
        creator.attach(self.env)

    @api.one
    def form_c0001ver02(self, creator):
        """
        GENERATE FORM ACCORDING TO form_c0001ver01
        """

        wb = creator.get_wb()

        # CREATE ALLOCATION SHEET
        # ----------------------------------------------------------
        create_allocation_sheet(self, wb)

        # WORK SHEET MAIN
        # ----------------------------------------------------------
        # Table 1
        row = 23
        column = 1
        sr = 1
        # Table 2
        #        row2 = 28
        #        column2 = 1

        row_count = len(self.invoice_ids) - 1

        signature_coor = "A32"
        logo_coor = "K2"
        header_coor = "A2"
        footer_row = 30
        ws = wb.get_sheet_by_name('main')

        ws.cell("A5").value = self.summary_no
        ws.cell("A7").value = fields.Datetime.from_string(self.create_date).strftime('%d-%b-%Y')
        ws.cell("A9").value = get_joined_value(self.mapped('invoice_ids.division_id.alias'))
        ws.cell("A11").value = get_joined_value(self.mapped('invoice_ids.cear_allocation_ids.cear_id.no'))
        ws.cell("A13").value = get_joined_value(
            self.mapped('invoice_ids.cear_allocation_ids.cear_id.authorized_amount'))

        cc_ac_list = []
        for oear in self.mapped('invoice_ids.oear_allocation_ids'):
            cc = oear.cost_center_id.cost_center
            ac = oear.account_code_id.account_code
            if cc and ac:
                cc_ac_list.append('-'.join([oear.cost_center_id.cost_center, oear.account_code_id.account_code]))
        ws.cell("A15").value = get_joined_value(cc_ac_list)

        # TODO CREATE LOGIC FOR CHARGE ACCOUNT
        ws.cell("A17").value = 'N/A'

        ws.cell("F5").value = get_joined_value(self.mapped('invoice_ids.contractor_id.name'))
        ws.cell("F7").value = get_joined_value(self.mapped('invoice_ids.contract_id.contract_ref'))
        ws.cell("F15").value = get_joined_value(self.mapped('invoice_ids.mms_no'))
        ws.cell("I7").value = get_joined_value(self.mapped('invoice_ids.po_id.no'))
        ws.cell("I11").value = get_joined_value(self.mapped('invoice_ids.po_id.amount'))

        # TODO CREATE LOGIC FOR CONTRACT VALIDITY
        ws.cell("F9").value = get_joined_value(self.mapped('invoice_ids.contract_id.end_date'))
        ws.cell("F11").value = get_joined_value(self.mapped('invoice_ids.contract_id.amount'))
        currency = self.mapped("invoice_ids.currency_id")
        ws.cell("F17").value = currency[0].name if currency else ""

        # Prepare Tables
        ws.insert_rows(row, row_count)  # Table 1
        # ws.insert_rows(row2 + row_count, row_count)  # Table 2

        # Populate Data
        invoices = self.invoice_ids.sorted(key=lambda rec: rec.sequence)
        # Table 1
        for r in invoices:
            ws.cell(row=row, column=column).value = sr
            ws.cell(row=row, column=column + 1).value = fields.Datetime.from_string(r.invoice_date).strftime('%d-%b-%Y')
            ws.cell(row=row, column=column + 2).value = r.invoice_no or ''
            ws.cell(row=row, column=column + 3).value = r.description or ''
            ws.cell(row=row, column=column + 4).value = r.region_id.alias.upper() or ''
            ws.cell(row=row, column=column + 5).value = '' if not r.due_percentage else r.due_percentage / 100
            ws.cell(row=row, column=column + 6).value = r.revenue_amount or ''
            ws.cell(row=row, column=column + 7).value = r.opex_amount or ''
            ws.cell(row=row, column=column + 8).value = r.capex_amount or ''
            ws.cell(row=row, column=column + 9).value = r.penalty_amount or ''
            ws.cell(row=row, column=column + 10).value = r.certified_invoice_amount or ''
            row += 1
            sr += 1
        # Table 2
        # sr = 1
        # row2 += row_count
        # for r in invoices:
        #     rfs_date = fields.Datetime.from_string(r.rfs_date)
        #     actual_rfs_date = fields.Datetime.from_string(r.actual_rfs_date)
        #     ws.cell(row=row2, column=column2).value = rfs_date.strftime('%d-%b-%Y')
        #     ws.cell(row=row2, column=column2 + 2).value = actual_rfs_date.strftime('%d-%b-%Y')
        #     ws.cell(row=row2, column=column2 + 3).value = '{} days'.format((actual_rfs_date - rfs_date).days)
        #     ws.cell(row=row2, column=column2 + 4).value = ''
        #     ws.cell(row=row2, column=column2 + 6).value = ''
        #     ws.merge_cells(start_row=row2, start_column=column2, end_row=row2, end_column=column2 + 1)
        #
        #     row2 += 1
        #     sr += 1

        # INSERT HEADER LOGO AND SIGNATURE
        inject_form_header(ws, creator, logo_coor, header_coor)

        # TABLE FORMAT
        for r in [28, 30]:  # A28, A30
            footer_cell = ws.cell(row=r + row_count, column=1)  # A28 + row_count - 2 row
            footer_cell.font = Font(size=11, bold=True)
            ws.row_dimensions[footer_cell.row].height = 83

        # FORMAT FOOTER
        format_footer(ws, footer_row, sr)

        # SIGNATURE
        inject_signature(ws, self.get_signatories(), signature_coor, sr)

        # SAVE FINAL ATTACHMENT
        creator.save()
        creator.attach(self.env)

    # ONCHANGE
    # ----------------------------------------------------------

    # BUTTONS/TRANSITIONS
    # ----------------------------------------------------------
    # TODO MUST DO A TRANSITION FOR RETURNING TO DRAFT

    @api.multi
    def set2draft(self):
        self.state = 'draft'

    @api.multi
    def set2file_generated(self):
        if len(self.invoice_ids) == 0:
            raise ValidationError('Empty Invoice List')

        currency_names = self.mapped("invoice_ids.currency_id.name")
        if len(set(currency_names)) > 1:
            raise ValidationError("Summary can only be generated for invoices with same currency.")

        creator = Creator(summary_no=self.summary_no,
                          summary_res_id=self.id,
                          form_filename=self.form)

        # get attribute from form name removing .xlsx
        # eg. self.form_c0001ver01
        form = getattr(self, self.form.split('.')[0])
        form(creator)

        # Set related invoices state to "summary generated"
        for invoice in self.mapped('invoice_ids'):
            invoice.state = 'summary generated'

        self.state = 'file generated'

    @api.multi
    def set2sd_signed(self):
        if self.sd_signed_date is False:
            self.sd_signed_date = fields.Date.today()

        for invoice in self.mapped('invoice_ids'):
            invoice.sd_signed_date = self.sd_signed_date
            invoice.state = 'sd signed'
        self.state = 'sd signed'

    @api.multi
    def set2svp_signed(self):
        if self.svp_signed_date is False:
            self.svp_signed_date = fields.Date.today()

        for invoice in self.mapped('invoice_ids'):
            invoice.svp_signed_date = self.svp_signed_date
            invoice.state = 'svp signed'
        self.state = 'svp signed'

    @api.multi
    def set2cto_signed(self):
        if self.cto_signed_date is False:
            self.cto_signed_date = fields.Date.today()

        for invoice in self.mapped('invoice_ids'):
            invoice.cto_signed_date = self.cto_signed_date
            invoice.state = 'cto signed'
        self.state = 'cto signed'

    @api.multi
    def set2sent_to_finance(self):
        if self.sent_finance_date is False:
            self.sent_finance_date = fields.Date.today()

        for invoice in self.mapped('invoice_ids'):
            invoice.sent_finance_date = self.sent_finance_date
            invoice.state = 'sent to finance'
        self.state = 'sent to finance'

    @api.multi
    def set2closed(self):
        if self.closed_date is False:
            self.closed_date = fields.Date.today()

        for invoice in self.mapped('invoice_ids'):
            if self.objective == 'on hold certification':
                invoice.release_hold_amount()
            invoice.closed_date = self.closed_date
            invoice.state = 'closed'
        self.state = 'closed'

    @api.multi
    def set2cancelled(self):
        for invoice in self.mapped('invoice_ids'):
            invoice.state = 'verified'
        self.state = 'cancelled'

    @api.multi
    def reset_summary(self):
        for invoice in self.mapped('invoice_ids'):
            invoice.state = 'verified'
        self.state = 'draft'

    # REDIRECT/OPEN OTHER VIEWS BUTTONS
    # ----------------------------------------------------------
    @api.multi
    def download_summary(self):
        attachment_id = self.env['ir.attachment'].search([
            ('res_model', '=', 'budget.invoice.invoice.summary'),
            ('res_id', '=', self.id)
        ], limit=1, order='id desc')

        # TODO CHECK HW TO REMOVE ?debug on the url
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment_id.id,
            'target': 'self',
            'res_id': self.id,
        }

    # MISC FUNCTIONS
    # ----------------------------------------------------------
    @api.multi
    def get_signatories(self):
        ids = self.mapped('signature_ids').ids
        signatories = self.signature_ids.create_signatories(ids)
        return signatories

    @api.multi
    def _delete_attachments(self):
        attachment = self.env['ir.attachment'].search([('res_model', '=', self._name)])
        attachment.unlink()
        return True

    # POLYMORPH FUNCTIONS
    # ----------------------------------------------------------
    @api.one
    def unlink(self):
        for invoice in self.mapped('invoice_ids'):
            invoice.state = 'verified'
        return super(InvoiceSummary, self).unlink()

    @api.model
    def create(self, vals):
        res = super(InvoiceSummary, self).create(vals)
        if self.env.context.get('auto_generate', False):
            res.set2file_generated()

        return res
